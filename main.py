import os
import math
import json
import random
import copy
import pandas as pd
from collections import defaultdict
from math import gcd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------
# Settings loader
# ----------------------------
def load_settings(path="settings.json"):
    DEFAULT = {
        "working_days": ["Mon", "Tue", "Wed", "Thu", "Fri"],
        "working_hours": ["9:00", "18:30"],
        "break_slots": ["12:30-13:30", "16:30-17:00"],
        "slot_durations": {"lec": 1.5, "lab": 2.0, "tut": 1.0}
    }
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                data = json.load(f)
        except Exception:
            data = DEFAULT
    else:
        data = DEFAULT
    breaks = []
    for b in data.get("break_slots", []):
        if isinstance(b, str) and "-" in b:
            a, c = b.split("-", 1)
            breaks.append((a.strip(), c.strip()))
        elif isinstance(b, (list, tuple)) and len(b) == 2:
            breaks.append((b[0], b[1]))
    data["break_slots"] = breaks
    return data

# ----------------------------
# Time helpers
# ----------------------------
def time_to_minutes(t):
    if isinstance(t, (int, float)):
        return int(t)
    parts = str(t).strip().split(":")
    if len(parts) == 1:
        h = int(parts[0]); m = 0
    else:
        h, m = map(int, parts)
    return h * 60 + m

def minutes_to_time(m):
    h = int(m // 60); mm = int(m % 60)
    return f"{h:02d}:{mm:02d}"

def gcd_list(nums):
    nums = [int(n) for n in nums if n and n > 0]
    if not nums:
        return 15
    g = nums[0]
    for n in nums[1:]:
        g = gcd(g, n)
    return g if g > 0 else 15

# ----------------------------
# Parsing helpers
# ----------------------------
def parse_list(cell):
    try:
        if pd.isna(cell) or str(cell).strip() == "":
            return []
    except Exception:
        pass
    return [x.strip().upper() for x in str(cell).split(",") if x.strip()]

def safe_upper(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return str(val).strip().upper()

# ----------------------------
# File reading helper
# ----------------------------
def read_input_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ----------------------------
# L-T-P parser
# ----------------------------
def parse_LTP_from_ltpsc(ltpsc):
    try:
        if pd.isna(ltpsc):
            return 0, 0, 0
    except Exception:
        pass
    s = str(ltpsc).strip()
    if not s:
        return 0, 0, 0
    parts = s.split("-")
    try:
        L = int(parts[0]) if len(parts) > 0 and parts[0] else 0
        T = int(parts[1]) if len(parts) > 1 and parts[1] else 0
        P = int(parts[2]) if len(parts) > 2 and parts[2] else 0
        return L, T, P
    except:
        return 0, 0, 0

# ----------------------------
# Build slot requests for one division
# ----------------------------
def build_slot_requests_for_division(df, div_fullname, settings):
    normals = []
    baskets = {}
    raw_rows = df.to_dict(orient="records")
    div_name_up = safe_upper(div_fullname)
    for idx, row in df.iterrows():
        elective_flag = safe_upper(row.get("ELECTIVE OR NOT", "NO"))
        sem_type = safe_upper(row.get("FULLSEM OR HALFSEM", "FULLSEM"))
        code = safe_upper(row.get("COURSE CODE", ""))
        title = safe_upper(row.get("COURSE TITLE", ""))
        faculty_list = parse_list(row.get("FACULTY", ""))
        class_asst = parse_list(row.get("CLASS ASSISTANTS", ""))
        lab_asst = parse_list(row.get("LAB ASSISTANTS", ""))
        ltpsc = row.get("L-T-P-S-C", row.get("L-T-P", ""))
        L, T, P = parse_LTP_from_ltpsc(ltpsc)
        room_no = parse_list(row.get("ROOM.NO", ""))
        lab_room_no = parse_list(row.get("LAB ROOM.NO", ""))
        slot_base = safe_upper(row.get("SLOT NAME", ""))
        merge_raw = row.get("MERGE", "")
        # split comma-separated list, strip spaces, upper-case
        merge_list = [safe_upper(m.strip()) for m in str(merge_raw).split(",") if m.strip()]
        # ensure self is included
        merge_set = set(merge_list)
        merge_set.add(div_name_up)
        merge_with = sorted(merge_set)

        # kinds: lec, tut, lab - compute occurrences using slot_durations
        for kind, hours in [("lec", L), ("tut", T), ("lab", P)]:
            if hours <= 0:
                continue
            kind_lower = kind.lower()
            dur_hours = settings["slot_durations"].get(kind_lower, 1.0)
            group_id = f"{code}__{kind_lower}__{slot_base}"
            
            total_hours = hours
            while total_hours > 0:
                dur = min(dur_hours, total_hours)
                dur_min = max(1, int(round(dur * 60)))
                slot_label = f"{slot_base}-{kind.upper()}"
                occ = {
                    "group_id": group_id,
                    "slot_base": slot_base,
                    "slot_label": slot_label,
                    "code": code,
                    "title": title,
                    "faculty": faculty_list,
                    "class_asst": class_asst,
                    "lab_asst": lab_asst,
                    "L": L, "T": T, "P": P,
                    "L-T-P-S-C": ltpsc,
                    "ROOM.NO": room_no,
                    "LAB ROOM.NO": lab_room_no,
                    "sem_type": sem_type,
                    "merge_with": merge_with,
                    "division": div_name_up,
                    "kind": kind_lower,
                    "_duration_hours": dur,
                    "_duration_min": dur_min
                }
                if elective_flag == "YES" and slot_base.lower().startswith("elective"):
                    basket_key = f"{slot_base}__{kind_lower}"
                    baskets.setdefault(basket_key, []).append(occ)
                else:
                    normals.append(occ)
                total_hours -= dur
    return normals, baskets, raw_rows

# Fix: trivial helper to avoid odd code line above
def kind_upper(x):
    return x.upper()

# ----------------------------
# Scheduling engine (minute-accurate; dynamic gap insertion)
# ----------------------------
def schedule_globally(all_normals_per_div, all_baskets, settings, min_gap_minutes, faculty_gap_minutes, max_attempts=20):
    days = settings["working_days"]
    wh_start = time_to_minutes(settings["working_hours"][0])
    wh_end = time_to_minutes(settings["working_hours"][1])
    dur_minutes = {k: int(v * 60) for k, v in settings["slot_durations"].items()}
    # Keep base_interval for initial header boundaries (we'll expand later)
    base_interval = gcd_list(list(dur_minutes.values()))
    if base_interval < 5:
        base_interval = 15

    # initial candidate start times (the original interval start points)
    interval_times = []
    t = wh_start
    while t < wh_end:
        interval_times.append(t)
        t += base_interval

    break_ranges = []
    for bstart, bend in settings.get("break_slots", []):
        bs = time_to_minutes(bstart); be = time_to_minutes(bend)
        break_ranges.append((bs, be))

    # Build master normal list and required counts per (group_id, division)
    normal_list_master = []
    required_per_div = defaultdict(int)
    for div, slots in all_normals_per_div.items():
        div_up = safe_upper(div)
        for s in slots:
            entry = s.copy()
            entry["_division"] = div_up
            duration_hours = entry.get("_duration_hours", settings["slot_durations"].get(entry.get("kind"), 1.0))
            entry["_duration_min"] = max(1, int(round(duration_hours * 60)))
            normal_list_master.append(entry)
            gid = entry.get("group_id")
            required_per_div[(gid, div_up)] += 1

    # Baskets -> synthetic gid
    baskets_master = {}
    for b_key, members in all_baskets.items():
        if not members:
            continue
        gid = f"BASKET__{b_key}"
        processed_members = []
        for m in members:
            mcopy = m.copy()
            mcopy["division"] = safe_upper(mcopy.get("division", ""))
            mcopy["_duration_min"] = max(1, int(round(mcopy.get("_duration_hours", 1.0) * 60)))
            processed_members.append(mcopy)
            required_per_div[(gid, mcopy["division"])] += 1
        baskets_master[gid] = processed_members

    # initialize best result
    best_result = None
    best_uns_count = None
    kind_priority = {"lec": 0, "tut": 1, "lab": 2}

    # We'll store placements now as minute-based entries:
    # placements[division][day] = list of dicts: {start_min, end_min, label, kind, meta}
    for attempt in range(max_attempts):
        random.seed(2000 + attempt)
        placements = {safe_upper(div): {d: [] for d in days} for div in all_normals_per_div.keys()}

        # per-person scheduled times (for faculty gap checks): person -> list of (day, start_min, end_min)
        occ_person_times = defaultdict(list)
        # per-division/day list of existing placements for overlap/room/person checks (we will use placements dict)
        placed_counts = defaultdict(int)

        normal_list = copy.deepcopy(normal_list_master)
        random.shuffle(normal_list)
        normal_list.sort(key=lambda x: (kind_priority.get(x["kind"], 3), -x["_duration_min"], random.random()))

        def overlaps(a_start, a_end, b_start, b_end):
            return not (a_end <= b_start or b_end <= a_start)

        def any_conflict_with_existing(merge_group, day, cand_start_min, cand_end_min, busy_people, rooms_set):
            # check breaks
            for bs, be in break_ranges:
                if not (cand_end_min <= bs or cand_start_min >= be):
                    return True
            # for each merged division, check existing placements
            for mdiv in merge_group:
                for ex in placements.get(mdiv, {}).get(day, []):
                    ex_s = ex["start_min"]; ex_e = ex["end_min"]
                    # if overlap, conflict
                    if overlaps(cand_start_min, cand_end_min, ex_s, ex_e):
                        return True
                    # enforce min_gap between placements of same course in same division
                    # note: we will check course slot min_gap rule separately where relevant
                    # but also ensure min_gap_minutes between unrelated placements is not mandatory (only for same course)
                # same-course min_gap: ensure new block is at least min_gap away from existing placements
                for ex in placements.get(mdiv, {}).get(day, []):
                    ex_s = ex["start_min"]; ex_e = ex["end_min"]
                    if not (cand_end_min + min_gap_minutes <= ex_s or cand_start_min >= ex_e + min_gap_minutes):
                        # This enforces the course min_gap rule across placements in same division
                        # but only if they are the same group_id we'll treat later; to be conservative we apply for same course/group
                        # We'll check same-course separately
                        pass
            # check rooms and people overlapping using placements entries
            for mdiv in merge_group:
                for ex in placements.get(mdiv, {}).get(day, []):
                    ex_meta = ex.get("meta", {})
                    ex_people = set()
                    ex_rooms = set()
                    if ex_meta:
                        if ex_meta.get("kind") in ("lec", "tut"):
                            ex_people.update(ex_meta.get("faculty", []) or [])
                            ex_people.update(ex_meta.get("class_asst", []) or [])
                            ex_rooms.update(ex_meta.get("ROOM.NO", []) or [])
                        else:
                            ex_people.update(ex_meta.get("faculty", []) or [])
                            ex_people.update(ex_meta.get("lab_asst", []) or [])
                            ex_rooms.update(ex_meta.get("LAB ROOM.NO", []) or [])
                    if busy_people & ex_people:
                        if overlaps(cand_start_min, cand_end_min, ex["start_min"], ex["end_min"]):
                            return True
                    if rooms_set and (rooms_set & ex_rooms):
                        if overlaps(cand_start_min, cand_end_min, ex["start_min"], ex["end_min"]):
                            return True
            # faculty gap global check
            for person in busy_people:
                for (pday, pstart, pend) in occ_person_times.get(person, []):
                    if pday != day:
                        continue
                    if not (cand_end_min + faculty_gap_minutes <= pstart or cand_start_min >= pend + faculty_gap_minutes):
                        return True
            return False

        def violates_same_course_day_rules(mdiv, day, group_id, kind, cand_start_min):
            # Prevent same course same kind twice a day etc (preserve original behaviour)
            for ex in placements.get(mdiv, {}).get(day, []):
                ex_kind = ex.get("kind")
                ex_meta = ex.get("meta", {})
                existing_gid = None
                if isinstance(ex_meta, dict):
                    existing_gid = ex_meta.get("group_id") or ex_meta.get("code")
                if not existing_gid:
                    continue
                if existing_gid == group_id:
                    # same kind twice
                    if ex_kind == kind:
                        return True
                    if set([ex_kind, kind]) == set(["lec", "lab"]):
                        ex_start_min = ex["start_min"]
                        # preserve previous logic: lab after lec ordering
                        if kind == "lab":
                            if ex_kind == "lec":
                                if cand_start_min <= ex_start_min:
                                    return True
                        elif ex_kind == "lab":
                            if cand_start_min >= ex_start_min:
                                return True
                    if "tut" in (ex_kind, kind):
                        return True
            return False

        def mark_placement_across_merged(merge_group, day, cand_start_min, cand_end_min, busy_people, rooms_set, meta, group_id, label, kind):
            # store placement dicts for each division in merge_group
            for mdiv in merge_group:
                placements[mdiv][day].append({
                    "start_min": cand_start_min,
                    "end_min": cand_end_min,
                    "label": label,
                    "kind": kind,
                    "meta": meta
                })
                if required_per_div.get((group_id, mdiv), 0) > 0:
                    if placed_counts.get((group_id, mdiv), 0) < required_per_div.get((group_id, mdiv), 0):
                        placed_counts[(group_id, mdiv)] += 1
            # mark person times
            for p in busy_people:
                occ_person_times[p].append((day, cand_start_min, cand_end_min))

        unscheduled = []

        # Place normal slots (minute-aware)
        for slot in normal_list:
            group_id = slot.get("group_id")
            merge_group_raw = slot.get("merge_with", []) or []
            if isinstance(merge_group_raw, str):
                merge_group_raw = [m.strip() for m in merge_group_raw.split(",") if m.strip()]
            merge_group = [safe_upper(m) for m in merge_group_raw if m]
            if not merge_group:
                merge_group = [slot["_division"]]
            # resolve names (preserve placements keys)
            resolved_merge = []
            for m in merge_group:
                if m in placements:
                    resolved_merge.append(m)
                else:
                    for pk in placements.keys():
                        if pk.replace(" ", "").upper() == m.replace(" ", "").upper():
                            resolved_merge.append(pk)
                            break
            if not resolved_merge:
                resolved_merge = [slot["_division"]]
            merge_group = resolved_merge

            # skip if already placed required count
            skip_flag = True
            for div in merge_group:
                if placed_counts.get((group_id, div), 0) < required_per_div.get((group_id, div), 0):
                    skip_flag = False; break
            if skip_flag:
                continue

            duration_min = slot.get("_duration_min", max(1, int(round(settings["slot_durations"].get(slot.get("kind"), 1.0) * 60))))
            busy_people = set(slot.get("faculty", []) or [])
            if slot.get("kind") in ("lec", "tut"):
                busy_people.update(slot.get("class_asst", []) or [])
                rooms = set(slot.get("ROOM.NO", []) or [])
            else:
                busy_people.update(slot.get("lab_asst", []) or [])
                rooms = set(slot.get("LAB ROOM.NO", []) or [])

            placed = False
            # days scored by current load
            day_scores = []
            for d in days:
                score = sum(len(placements.get(div, {}).get(d, [])) for div in merge_group)
                day_scores.append((score, d))
            random.shuffle(day_scores)
            day_scores.sort(key=lambda x: x[0])

            for _, day in day_scores:
                # candidate start times: original interval_times (minute aligned)
                start_candidates = list(interval_times)
                random.shuffle(start_candidates)
                start_candidates.sort()
                for cand in start_candidates:
                    # initial candidate start and end
                    cand_start_min = cand
                    cand_end_min = cand_start_min + duration_min
                    # ensure block fits within working hours
                    if cand_end_min > wh_end:
                        continue
                    # skip if overlaps break
                    bad = False
                    for bs, be in break_ranges:
                        if not (cand_end_min <= bs or cand_start_min >= be):
                            bad = True; break
                    if bad:
                        continue
                    # same-course/day rules
                    violated = False
                    for mdiv in merge_group:
                        if violates_same_course_day_rules(mdiv, day, group_id, slot.get("kind"), cand_start_min):
                            violated = True; break
                    if violated:
                        continue

                    # Now check if there's a placement immediately before cand_start_min in any merged division
                    # If previous placement ends exactly at cand_start_min (or ends within min_gap), we attempt to shift the candidate forward by min_gap_minutes
                    need_shift = False
                    for mdiv in merge_group:
                        for ex in placements.get(mdiv, {}).get(day, []):
                            ex_end = ex["end_min"]
                            # if ex_end == cand_start_min OR ex_end + min_gap_minutes > cand_start_min (i.e. too close), we prefer shifting
                            if ex_end == cand_start_min or (0 <= (cand_start_min - ex_end) < min_gap_minutes):
                                need_shift = True
                                break
                        if need_shift:
                            break

                    if need_shift:
                        shifted_start = cand_start_min + min_gap_minutes
                        shifted_end = shifted_start + duration_min
                        # ensure shifted fits
                        if shifted_end > wh_end:
                            continue
                        # ensure not overlapping break
                        bad2 = False
                        for bs, be in break_ranges:
                            if not (shifted_end <= bs or shifted_start >= be):
                                bad2 = True; break
                        if bad2:
                            continue
                        # check conflicts with existing placements (rooms/people/faculty gap) on shifted interval
                        if any_conflict_with_existing(merge_group, day, shifted_start, shifted_end, busy_people, rooms):
                            continue
                        # also check same-course/day rules for shifted start
                        violated2 = False
                        for mdiv in merge_group:
                            if violates_same_course_day_rules(mdiv, day, group_id, slot.get("kind"), shifted_start):
                                violated2 = True; break
                        if violated2:
                            continue
                        # OK — mark placement at shifted times
                        label = slot.get("slot_label")
                        mark_placement_across_merged(merge_group, day, shifted_start, shifted_end, busy_people, rooms, slot, group_id, label, slot.get("kind"))
                        placed = True
                        break

                    else:
                        # no need to shift — check conflicts at original cand times
                        if any_conflict_with_existing(merge_group, day, cand_start_min, cand_end_min, busy_people, rooms):
                            continue
                        # OK mark placement
                        label = slot.get("slot_label")
                        mark_placement_across_merged(merge_group, day, cand_start_min, cand_end_min, busy_people, rooms, slot, group_id, label, slot.get("kind"))
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                unscheduled.append(slot)

        # Place baskets (electives grouped) — simplified minute-aware placement
        for b_key, members in baskets_master.items():
            # group members per division
            div_to_members = defaultdict(list)
            for m in members:
                merge_group = m.get("merge_with", []) or [m.get("division", "")]
                for div in merge_group:
                    div_up = safe_upper(div)
                    div_to_members[div_up].append(m)
            basket_divs = list(div_to_members.keys())

            combined_people = set()
            combined_rooms = set()
            max_duration_min = 0
            kind = members[0].get("kind", "lec")
            slot_base = members[0].get("slot_base", "")
            for m in members:
                combined_people.update(m.get("faculty", []) or [])
                if kind in ("lec", "tut"):
                    combined_people.update(m.get("class_asst", []) or [])
                    combined_rooms.update(m.get("ROOM.NO", []) or [])
                else:
                    combined_people.update(m.get("lab_asst", []) or [])
                    combined_rooms.update(m.get("LAB ROOM.NO", []) or [])
                duration_min = m.get("_duration_min", dur_minutes.get(kind, 60))
                max_duration_min = max(max_duration_min, duration_min)
            duration_min = max_duration_min

            placed = False
            day_scores = []
            for d in days:
                score = sum(len(placements.get(div, {}).get(d, [])) for div in basket_divs)
                day_scores.append((score, d))
            random.shuffle(day_scores)
            day_scores.sort(key=lambda x: x[0])

            for _, day in day_scores:
                for cand in interval_times:
                    cand_start_min = cand
                    cand_end_min = cand_start_min + duration_min
                    if cand_end_min > wh_end:
                        continue
                    # skip if overlaps break
                    bad = False
                    for bs, be in break_ranges:
                        if not (cand_end_min <= bs or cand_start_min >= be):
                            bad = True; break
                    if bad:
                        continue
                    # try shifting if immediate previous ends at cand_start_min
                    need_shift = False
                    for div in basket_divs:
                        for ex in placements.get(div, {}).get(day, []):
                            if ex["end_min"] == cand_start_min or (0 <= (cand_start_min - ex["end_min"]) < min_gap_minutes):
                                need_shift = True
                                break
                        if need_shift:
                            break
                    if need_shift:
                        shifted_start = cand_start_min + min_gap_minutes
                        shifted_end = shifted_start + duration_min
                        if shifted_end > wh_end:
                            continue
                        # skip breaks
                        bad2 = False
                        for bs, be in break_ranges:
                            if not (shifted_end <= bs or shifted_start >= be):
                                bad2 = True; break
                        if bad2:
                            continue
                        if any_conflict_with_existing(basket_divs, day, shifted_start, shifted_end, combined_people, combined_rooms):
                            continue
                        # OK place
                        for div in basket_divs:
                            placements[div][day].append({
                                "start_min": shifted_start,
                                "end_min": shifted_end,
                                "label": f"{slot_base}-{kind.upper()}",
                                "kind": kind,
                                "meta": {"basket_members": members, "slot_base": slot_base, "group_id": b_key, "faculty": list(combined_people), "ROOM.NO": list(combined_rooms)}
                            })
                            if placed_counts.get((b_key, div), 0) < required_per_div.get((b_key, div), 0):
                                placed_counts[(b_key, div)] += 1
                        for p in combined_people:
                            occ_person_times[p].append((day, shifted_start, shifted_end))
                        placed = True
                        break
                    else:
                        if any_conflict_with_existing(basket_divs, day, cand_start_min, cand_end_min, combined_people, combined_rooms):
                            continue
                        # OK place
                        for div in basket_divs:
                            placements[div][day].append({
                                "start_min": cand_start_min,
                                "end_min": cand_end_min,
                                "label": f"{slot_base}-{kind.upper()}",
                                "kind": kind,
                                "meta": {"basket_members": members, "slot_base": slot_base, "group_id": b_key, "faculty": list(combined_people), "ROOM.NO": list(combined_rooms)}
                            })
                            if placed_counts.get((b_key, div), 0) < required_per_div.get((b_key, div), 0):
                                placed_counts[(b_key, div)] += 1
                        for p in combined_people:
                            occ_person_times[p].append((day, cand_start_min, cand_end_min))
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                unscheduled.append({"basket_label": b_key})

        uns_count = len(unscheduled)
        if best_uns_count is None or uns_count < best_uns_count:
            best_uns_count = uns_count
            best_result = (copy.deepcopy(placements), [u for u in unscheduled], interval_times, base_interval, break_ranges)
        if uns_count == 0:
            break

    if best_result is None:
        placements = {safe_upper(div): {d: [] for d in days} for div in all_normals_per_div.keys()}
        return placements, ["Scheduling failed (no valid attempt)"], interval_times, base_interval, break_ranges

    return best_result
# ----------------------------
# Excel utilities (minute-aware)
# ----------------------------
def set_value_in_merged_region(ws, row, col_start, col_end, value):
    unmerge_ranges_overlapping(ws, row, col_start, col_end)
    if col_end > col_start:
        try:
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        except Exception:
            pass
    try:
        ws.cell(row=row, column=col_start, value=value)
    except Exception:
        try:
            ws.unmerge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        except Exception:
            pass
        try:
            ws.cell(row=row, column=col_start, value=value)
        except Exception:
            pass

def safe_sheet_title(raw):
    try:
        if pd.isna(raw): return None
    except Exception:
        pass
    s = str(raw) if raw is not None else ""
    s = s.strip()
    if s == "" or s.lower() == "nan": return None
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        s = s.replace(ch, "_")
    if len(s) > 31: s = s[:31]
    return s

def ranges_overlap(a_start, a_end, b_start, b_end):
    return not (a_end < b_start or b_end < a_start)

def unmerge_ranges_overlapping(ws, row, col_start, col_end):
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.min_col, mr.min_row, mr.max_col, mr.max_row
        if row >= min_row and row <= max_row:
            if ranges_overlap(col_start, col_end, min_col, max_col):
                to_unmerge.append(str(mr))
    for rng in to_unmerge:
        try:
            ws.unmerge_cells(rng)
        except:
            pass

# ----------------------------
# Build unallotted rows helper
# ----------------------------
def build_unallotted_rows(unscheduled_list, baskets_map):
    rows = []
    for u in unscheduled_list:
        if isinstance(u, dict) and "basket_label" in u:
            bkey = u["basket_label"]
            members = baskets_map.get(bkey, [])
            if members:
                for m in members:
                    div = safe_upper(m.get("division", ""))
                    slot_name = m.get("slot_base", "") or ""
                    code = m.get("code", "") or ""
                    title = m.get("title", "") or ""
                    kind = (m.get("kind") or "").upper()
                    faculty = ", ".join(m.get("faculty") or [])
                    class_ass = ", ".join(m.get("class_asst") or [])
                    lab_ass = ", ".join(m.get("lab_asst") or [])
                    rooms = ", ".join(m.get("ROOM.NO") or [])
                    labrooms = ", ".join(m.get("LAB ROOM.NO") or [])
                    merge = ", ".join(m.get("merge_with") or [])
                    rows.append({
                        "DIVISION": div,
                        "SLOT NAME": slot_name,
                        "COURSE CODE": code,
                        "COURSE NAME": title,
                        "KIND": kind,
                        "FACULTY": faculty,
                        "CLASS ASSISTANTS": class_ass,
                        "LAB ASSISTANTS": lab_ass,
                        "ROOM.NO": rooms,
                        "LAB ROOM.NO": labrooms,
                        "MERGE": merge,
                        "REASON": "NO VALID SLOT (BASKET)"
                    })
            else:
                rows.append({
                    "DIVISION": "",
                    "SLOT NAME": "",
                    "COURSE CODE": "",
                    "COURSE NAME": "",
                    "KIND": "",
                    "FACULTY": "",
                    "CLASS ASSISTANTS": "",
                    "LAB ASSISTANTS": "",
                    "ROOM.NO": "",
                    "LAB ROOM.NO": "",
                    "MERGE": bkey,
                    "REASON": "NO VALID SLOT (BASKET, MEMBERS UNKNOWN)"
                })
        elif isinstance(u, dict):
            div = safe_upper(u.get("division", u.get("_division", "")))
            slot_name = u.get("slot_base", "") or ""
            code = u.get("code", "") or ""
            title = u.get("title", "") or ""
            kind = (u.get("kind") or "").upper()
            faculty = ", ".join(u.get("faculty") or [])
            class_ass = ", ".join(u.get("class_asst") or [])
            lab_ass = ", ".join(u.get("lab_asst") or [])
            rooms = ", ".join(u.get("ROOM.NO") or [])
            labrooms = ", ".join(u.get("LAB ROOM.NO") or [])
            merge = ", ".join(u.get("merge_with") or [])
            rows.append({
                "DIVISION": div,
                "SLOT NAME": slot_name,
                "COURSE CODE": code,
                "COURSE NAME": title,
                "KIND": kind,
                "FACULTY": faculty,
                "CLASS ASSISTANTS": class_ass,
                "LAB ASSISTANTS": lab_ass,
                "ROOM.NO": rooms,
                "LAB ROOM.NO": labrooms,
                "MERGE": merge,
                "REASON": "NO VALID SLOT"
            })
        else:
            rows.append({
                "DIVISION": "",
                "SLOT NAME": "",
                "COURSE CODE": "",
                "COURSE NAME": str(u),
                "KIND": "",
                "FACULTY": "",
                "CLASS ASSISTANTS": "",
                "LAB ASSISTANTS": "",
                "ROOM.NO": "",
                "LAB ROOM.NO": "",
                "MERGE": "",
                "REASON": "NO VALID SLOT (UNKNOWN ITEM)"
            })
    return rows

# ----------------------------
# Write Excel (minute-aware header generation)
# ----------------------------
def write_year_excel(year, half_tag, placements, initial_interval_times, base_interval, break_ranges, colors, course_info_rows_per_div, settings, outdir=None, unallotted_rows=None):
    if outdir is None:
        outdir = os.path.join("timetable_outputs", f"Year_{year}")
    os.makedirs(outdir, exist_ok=True)
    fname = os.path.join(outdir, f"Timetable_Year{year}_{half_tag}.xlsx")
    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass
    days = settings["working_days"]

    # Build a set of boundaries: include original interval boundaries and all placement start/end and break boundaries
    boundaries = set()
    wh_start = time_to_minutes(settings["working_hours"][0])
    wh_end = time_to_minutes(settings["working_hours"][1])
    # include original coarse intervals (start & end)
    for t in initial_interval_times:
        boundaries.add(t)
        boundaries.add(t + base_interval)
    # include breaks boundaries
    for bs, be in break_ranges:
        boundaries.add(bs); boundaries.add(be)
    # include all placement boundaries
    for div, day_map in placements.items():
        for day, plist in day_map.items():
            for p in plist:
                boundaries.add(p["start_min"])
                boundaries.add(p["end_min"])
    # clamp boundaries to working hours and create sorted list
    boundaries = sorted([b for b in boundaries if b >= wh_start and b <= wh_end])
    # if boundaries does not include start or end, add them
    if wh_start not in boundaries:
        boundaries.insert(0, wh_start)
    if wh_end not in boundaries:
        boundaries.append(wh_end)
    # create intervals from consecutive boundaries
    time_intervals = []
    for i in range(len(boundaries) - 1):
        s = boundaries[i]; e = boundaries[i+1]
        if e > s:
            time_intervals.append((s, e))

    # Build header strings for each computed interval
    time_headers = [f"{minutes_to_time(s)} - {minutes_to_time(e)}" for s, e in time_intervals]

    # For each division, create sheet and fill the grid
    for div_index, (div, day_map) in enumerate(placements.items(), start=1):
        title_candidate = safe_sheet_title(div)
        if title_candidate is None:
            title_candidate = f"Div_{div_index}"
        base_title = title_candidate
        suffix = 1
        while title_candidate in wb.sheetnames:
            title_candidate = f"{base_title[:28]}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(title=title_candidate)
        ws.append([f"Division: {div}    Year: {year}    Half: {half_tag}"])
        ws.append([])
        header = ["Day/Time"] + time_headers
        ws.append(header)
        header_row_idx = ws.max_row
        # append rows for days
        for day in days:
            ws.append([day] + [""] * len(time_intervals))
        first_day_row = header_row_idx + 1

        # fill placements into grid (we will fill cell by cell; gaps will be left blank as per your Option B)
        # helper: find placement covering a given interval cell
        def find_placement_covering(plist, cell_start, cell_end):
            for p in plist:
                # placement covers this cell if [cell_start, cell_end) is within [p.start, p.end)
                if not (cell_end <= p["start_min"] or cell_start >= p["end_min"]):
                    # We want to assign label only if the cell is mostly within the placement
                    # We'll return the placement if overlap exists (this keeps cells representing parts of the class)
                    return p
            return None

        for r_idx, day in enumerate(days):
            plist = day_map.get(day, [])
            # for each interval cell
            for c_idx, (cell_s, cell_e) in enumerate(time_intervals):
                excel_row = first_day_row + r_idx
                excel_col = 2 + c_idx
                p = find_placement_covering(plist, cell_s, cell_e)
                if p:
                    # If this cell is a GAP interval (length == min_gap and there is no placement exactly covering it), we leave blank (Option B)
                    # But we want to show class cells: write label for part of class
                    # To avoid repeated writing many times, we'll write label at start boundary of that placement
                    # Find if this cell is the left-most cell inside that placement for this day/div
                    is_leftmost = (cell_s == p["start_min"])
                    if is_leftmost:
                        # compute how many consecutive interval cells belong to this same placement (for merging)
                        span = 0
                        cc = c_idx
                        while cc < len(time_intervals):
                            cs, ce = time_intervals[cc]
                            # if this interval overlaps with p
                            if not (ce <= p["start_min"] or cs >= p["end_min"]):
                                span += 1
                                cc += 1
                            else:
                                break
                        label = p.get("label") or (p.get("meta", {}).get("slot_base","") + "-" + (p.get("kind","")).upper() if isinstance(p.get("meta",{}), dict) else "")
                        col_start = excel_col
                        col_end = excel_col + span - 1
                        set_value_in_merged_region(ws, excel_row, col_start, col_end, label)
                        # style cells
                        slot_base = None
                        meta = p.get("meta", {})
                        if isinstance(meta, dict):
                            slot_base = meta.get("slot_base")
                        if not slot_base:
                            if isinstance(label, str) and "-" in label:
                                slot_base = "-".join(label.split("-")[:-1])
                            else:
                                slot_base = label
                        if slot_base in colors:
                            fill_color = colors[slot_base]
                        else:
                            fill_color = "#" + "".join(random.choices("0123456789ABCDEF", k=6))
                            colors[slot_base] = fill_color
                        color_code = fill_color[1:] if str(fill_color).startswith("#") else str(fill_color)
                        if not (isinstance(color_code, str) and len(color_code) in (6,8) and all(ch in "0123456789ABCDEFabcdef" for ch in color_code)):
                            color_code = "DDDDDD"
                        for cc in range(col_start, col_end + 1):
                            ccell = ws.cell(row=excel_row, column=cc)
                            ccell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                            ccell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            ccell.font = Font(size=10, bold=True)
                    else:
                        # cell inside the placement but not leftmost: we'll have been merged, so skip direct writes
                        pass
                else:
                    # No placement overlapping this cell.
                    # If this cell corresponds to a break interval, mark as BREAK
                    for bs, be in break_ranges:
                        if cell_s >= bs and cell_e <= be:
                            ccell = ws.cell(row=excel_row, column=excel_col)
                            ccell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                            if not ccell.value:
                                ccell.value = "BREAK"
                                ccell.alignment = Alignment(horizontal="center", vertical="center")
                            break
                    # If it's a tiny gap interval (equal to min_gap and not covered by any placement), leave blank (Option B)

        # column widths
        ws.column_dimensions[get_column_letter(1)].width = 14
        for ci in range(2, 2 + len(time_intervals)):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        # reference table
        ws.append([])
        ws.append(["Reference Table"])
        ref_columns = ["SLOT NAME", "FULLSEM OR HALFSEM", "COURSE CODE", "COURSE TITLE", "FACULTY", "CLASS ASSISTANTS", "LAB ASSISTANTS", "L-T-P-S-C", "ROOM.NO", "LAB ROOM.NO"]
        ws.append(ref_columns)
        raw_rows = course_info_rows_per_div.get(div, [])
        for r in raw_rows:
            row_values = []
            for col in ref_columns:
                if col == "FULLSEM OR HALFSEM":
                    val = r.get("FULLSEM OR HALFSEM", "")
                else:
                    val = r.get(col, "")
                row_values.append(val)
            ws.append(row_values)

    # Unallotted
    if unallotted_rows:
        try:
            ws_un = wb.create_sheet(title="Unallotted Slots")
        except Exception:
            suffix = 1
            name_try = "Unallotted Slots"
            while name_try in wb.sheetnames:
                name_try = f"Unallotted Slots_{suffix}"
                suffix += 1
            ws_un = wb.create_sheet(title=name_try)
        un_cols = ["DIVISION", "SLOT NAME", "COURSE CODE", "COURSE NAME", "KIND", "FACULTY", "CLASS ASSISTANTS", "LAB ASSISTANTS", "ROOM.NO", "LAB ROOM.NO", "MERGE", "REASON"]
        ws_un.append(un_cols)
        for ur in unallotted_rows:
            row_vals = [ur.get(c, "") for c in un_cols]
            ws_un.append(row_vals)

    if not wb.sheetnames:
        wb.create_sheet(title="Timetable")
    wb.save(fname)
    print(f"Saved: {fname}")

# ----------------------------
# Main program
# ----------------------------
def main():
    settings = load_settings("settings.json")
    print("Timetable Generator (improved: multi-value & merge-aware, stricter conflict checks)")
    print("-" * 70)
    print("Working days:", ", ".join(settings["working_days"]))
    wh = settings["working_hours"]
    print("Working hours:", f"{wh[0]} - {wh[1]}")
    breaks = settings.get("break_slots", [])
    if breaks:
        print("Break slots:", ", ".join([f"{b[0]}-{b[1]}" for b in breaks]))
    else:
        print("Break slots: None")
    print("\nSlot durations:")
    sd = settings.get("slot_durations", {})
    print(f"  Lecture (lec): {sd.get('lec', 1.0)} hours")
    print(f"  Lab     (lab): {sd.get('lab', 1.0)} hours")
    print(f"  Tutorial(tut): {sd.get('tut', 1.0)} hours")
    print("-" * 70)

    DEFAULT_MIN_GAP = 5
    DEFAULT_FACULTY_GAP = 180

    while True:
        try:
            raw = input(f"Enter minimum gap between consecutive slots in minutes (default {DEFAULT_MIN_GAP}): ") or str(DEFAULT_MIN_GAP)
            min_gap = int(raw)
            if min_gap < 0:
                print("Please enter a non-negative integer"); continue
            break
        except Exception:
            print("Please enter integer minutes")
    while True:
        try:
            raw = input(f"Enter minimum gap required for faculty between classes in minutes (default {DEFAULT_FACULTY_GAP}): ") or str(DEFAULT_FACULTY_GAP)
            faculty_gap = int(raw)
            if faculty_gap < 0:
                print("Please enter a non-negative integer"); continue
            break
        except Exception:
            print("Please enter integer minutes")
    print("Minimum gap (course slots):", min_gap, "minutes")
    print("Minimum gap (faculty):", faculty_gap, "minutes")
    print("-" * 40)

    n_years = int(input("Enter number of academic years: ").strip())
    inputs_per_year = {}
    for y in range(1, n_years + 1):
        n_div = int(input(f"Year {y}, number of divisions: ").strip())
        inputs_per_year[y] = {}
        for d in range(1, n_div + 1):
            full_div = input(f"  Short name for Division {d} (use exact name for MERGE, e.g. '1CSEA'): ").strip()
            if not full_div:
                full_div = input("    Division name cannot be blank — please enter short division name: ").strip()
            path = input(f"     Path to Excel/CSV for {full_div}: ").strip()
            inputs_per_year[y][full_div] = path

    for y in range(1, n_years + 1):
        print(f"\nProcessing Year {y} ...")
        div_paths = inputs_per_year[y]
        normals_first = {}
        normals_second = {}
        baskets_first = {}
        baskets_second = {}
        course_info_rows = {}
        slot_bases_set = set()
        for div_full, path in div_paths.items():
            div_up = safe_upper(div_full)
            if not os.path.exists(path):
                print(f" File not found: {path} for {div_full} — skipping division")
                normals_first[div_up] = []
                normals_second[div_up] = []
                course_info_rows[div_up] = []
                continue
            df = read_input_file(path)
            rows = df.to_dict(orient='records')
            course_info_rows[div_up] = rows
            normals, baskets, _ = build_slot_requests_for_division(df, div_full, settings)
            normals_f = [n for n in normals if safe_upper(n.get("sem_type", "FULLSEM")) in ("FULLSEM", "HALFSEM-1")]
            normals_s = [n for n in normals if safe_upper(n.get("sem_type", "FULLSEM")) in ("FULLSEM", "HALFSEM-2")]
            normals_first[div_up] = normals_f
            normals_second[div_up] = normals_s
            for b_key, members in baskets.items():
                sems = [safe_upper(m.get("sem_type", "FULLSEM")) for m in members]
                if any(s in ("FULLSEM", "HALFSEM-1") for s in sems):
                    baskets_first.setdefault(b_key, []).extend(members)
                if any(s in ("FULLSEM", "HALFSEM-2") for s in sems):
                    baskets_second.setdefault(b_key, []).extend(members)
            for n in normals:
                sb = n.get("slot_base") or ""
                if sb:
                    slot_bases_set.add(sb)

        # deterministic colors
        colors = {}
        palette = [
            "FF5733", "FF8D1A", "FFC300", "FFEA00", "9AFB60", "2ECC71", "27AE60",
            "00B2FF", "3498DB", "6C5CE7", "9B59B6", "F06292", "FFB6C1", "FF7F50",
            "D35400", "E67E22", "F39C12", "F1C40F", "1ABC9C", "16A085"
        ]
        slot_bases_sorted = sorted(list(slot_bases_set))
        for i, s in enumerate(slot_bases_sorted):
            if i < len(palette):
                colors[s] = "#" + palette[i]
            else:
                while True:
                    rnd_color = "#" + "".join(random.choices("0123456789ABCDEF", k=6))
                    if rnd_color not in colors.values():
                        colors[s] = rnd_color
                        break

        placements_first, uns_first, interval_times, base_interval, break_ranges = schedule_globally(normals_first, baskets_first, settings, min_gap, faculty_gap)
        unallotted_rows_first = build_unallotted_rows(uns_first if isinstance(uns_first, list) else [], baskets_first)
        write_year_excel(y, "first_halfsem", placements_first, interval_times, base_interval, break_ranges, colors, course_info_rows, settings, unallotted_rows=unallotted_rows_first)

        placements_second, uns_second, interval_times2, base_interval2, break_ranges2 = schedule_globally(normals_second, baskets_second, settings, min_gap, faculty_gap)
        unallotted_rows_second = build_unallotted_rows(uns_second if isinstance(uns_second, list) else [], baskets_second)
        write_year_excel(y, "second_halfsem", placements_second, interval_times2, base_interval2, break_ranges2, colors, course_info_rows, settings, unallotted_rows=unallotted_rows_second)

        uns_total = []
        if isinstance(uns_first, list):
            uns_total.extend([str(u) for u in uns_first])
        if isinstance(uns_second, list):
            uns_total.extend([str(u) for u in uns_second])
        if uns_total:
            print("\n Unscheduled items (may need input adjustments):")
            for u in uns_total[:200]:
                print("  ", u)
            if len(uns_total) > 200:
                print("   ...", len(uns_total) - 200, "more not shown ...")

    print("\nAll done. Timetables saved in ./timetable_outputs")

if __name__ == "__main__":
    main()
