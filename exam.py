# seating_scheduler_final_seating_sessions.py
import os
import math
import random
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

random.seed(42)

# -------------------------
# Helpers
# -------------------------
def safe_int(v):
    try:
        s = str(v).strip()
        if s.lower() in ("", "nan", "none"):
            return 0
        return int(float(s))
    except:
        return 0

def base_slotname(slot_name):
    s = str(slot_name).strip().upper()
    if "_Y" in s:
        return s.split("_Y")[0]
    if "_" in s:
        return s.split("_")[0]
    return s

def inv_key(num, name):
    return f"{num}|{name}"

def inv_display_from_key(key):
    try:
        num, name = key.split("|", 1)
        return f"{num} - {name}"
    except:
        return str(key)

# -------------------------
# Hardcoded inputs (as requested)
# -------------------------
divisions = {
    1: {"1CSEA": r"project\1CSEA.xlsx",
        "1CSEB": r"project\1CSEB.xlsx",
        "1DSAI": r"project\1DSAI.xlsx",
        "1ECE": r"project\1ECE.xlsx"},
    2: {"2CSEA": r"project\2CSEA.xlsx",
        "2CSEB": r"project\2CSEB.xlsx",
        "2DSAI": r"project\2DSAI.xlsx",
        "2ECE": r"project\2ECE.xlsx"},
    3: {"3CSEA": r"project\3CSEA.xlsx",
        "3CSEB": r"project\3CSEB.xlsx",
        "3DSAI": r"project\3DSAI.xlsx",
        "3ECE": r"project\3ECE.xlsx"},
    4: {"4CSEA": r"project\4CSEA.xlsx",
        "4CSEB": r"project\4CSEB.xlsx",
        "4DSAI": r"project\4DSAI.xlsx",
        "4ECE": r"project\4ECE.xlsx"}
}

rooms_path = r"project\Rooms.xlsx"
invig_path = r"project\invigilators_list.xlsx"
num_years = 4

# -------------------------
# Load courses
# -------------------------
def load_courses(divisions_dict):
    rows = []
    for year, divs in divisions_dict.items():
        for div, path in divs.items():
            df = pd.read_excel(path, engine="openpyxl")
            df.columns = [str(c).strip() for c in df.columns]
            for _, r in df.iterrows():
                elective = str(r.get("ELECTIVE OR NOT", "")).strip().upper()
                half_type = str(r.get("FULLSEM OR HALFSEM", "")).strip().upper()
                course_code = str(r.get("COURSE CODE", "")).strip()
                course_title = str(r.get("COURSE TITLE", "")).strip()
                slot_raw = str(r.get("SLOT NAME", "")).strip().upper()
                merge_raw = str(r.get("MERGE", "")).strip().upper()
                no_students = safe_int(r.get("NO. OF STUDENTS", 0))

                merge_list = [m.strip().upper() for m in merge_raw.split(",") if str(m).strip()]
                if div not in merge_list:
                    merge_list.append(div)

                slot_name = slot_raw
                if elective == "YES":
                    slot_name = f"{slot_raw}_Y{year}"

                rows.append({
                    "YEAR": year,
                    "DIVISION": div,
                    "ELECTIVE": elective,
                    "FULLSEM_TYPE": half_type,
                    "SLOT": slot_name,
                    "SLOT_RAW": slot_raw,
                    "COURSE_CODE": course_code,
                    "COURSE_TITLE": course_title,
                    "MERGE": merge_list,
                    "NO_STUDENTS": no_students
                })
    df_all = pd.DataFrame(rows)
    for c in ["DIVISION", "SLOT", "SLOT_RAW"]:
        if c in df_all.columns:
            df_all[c] = df_all[c].astype(str).str.upper()
    return df_all

# -------------------------
# Split halves
# -------------------------
def split_half(df):
    first = df[df["FULLSEM_TYPE"].isin(["FULLSEM", "HALFSEM-1"])].copy().reset_index(drop=True)
    second = df[df["FULLSEM_TYPE"].isin(["FULLSEM", "HALFSEM-2"])].copy().reset_index(drop=True)
    return first, second

# -------------------------
# Allocate slots by seating capacity
# -------------------------
def allocate_slots_by_seating_capacity(courses_df, rooms_df):
    slot_map = {}
    for _, r in courses_df.iterrows():
        key = r["SLOT"]
        if key not in slot_map:
            slot_map[key] = {"slot_key": key, "slot_raw": r.get("SLOT_RAW", key), "courses": [], "divisions": set(), "merged_flag": False}
        slot_map[key]["courses"].append(r)
        slot_map[key]["divisions"].add(r["DIVISION"])
        if isinstance(r["MERGE"], (list, tuple)) and len(r["MERGE"]) > 1:
            slot_map[key]["merged_flag"] = True

    slots = []
    for k, v in slot_map.items():
        if v["merged_flag"]:
            students = max(int(rr["NO_STUDENTS"]) for rr in v["courses"])
        else:
            students = sum(int(rr["NO_STUDENTS"]) for rr in v["courses"])
        slots.append({
            "slot_key": k,
            "slot_raw": v["slot_raw"],
            "courses": v["courses"],
            "divisions": v["divisions"],
            "merged_flag": v["merged_flag"],
            "students": students,
            "assigned": False
        })

    total_students = sum(s["students"] for s in slots)
    rooms_sorted = rooms_df.sort_values(by="Seating Capacity", ascending=False)
    room_caps = [int(r["Seating Capacity"]) for _, r in rooms_sorted.iterrows()]
    session_capacity = sum([cap // 2 for cap in room_caps])
    min_days = math.ceil(total_students / session_capacity) if session_capacity > 0 else 1
    sessions_per_day = [{"FN": True, "AN": True} for _ in range(min_days)]
    last_day_total = total_students - session_capacity*(min_days-1)
    if last_day_total <= max(room_caps)//2:
        sessions_per_day[-1] = {"FN": True, "AN": False}

    assignments = []
    division_taken = defaultdict(lambda: defaultdict(bool))
    slot_order = sorted(slots, key=lambda x: x["slot_key"])
    day_idx = 1
    for day_i in range(min_days):
        for sess in ("FN", "AN"):
            if not sessions_per_day[day_i][sess]:
                continue
            remaining_capacity = session_capacity
            placed_slots = []
            for slot in slot_order:
                if slot["assigned"]:
                    continue
                if slot["students"] > session_capacity:
                    continue
                conflict = any(division_taken[(day_idx, sess)].get(div, False) for div in slot["divisions"])
                if conflict:
                    continue
                if slot["students"] <= remaining_capacity:
                    placed_slots.append(slot)
                    remaining_capacity -= slot["students"]
                    slot["assigned"] = True
                    for div in slot["divisions"]:
                        division_taken[(day_idx, sess)][div] = True
            # append even if placed_slots empty to preserve day/session structure? original code appended regardless.
            assignments.append({"day": day_idx, "session": sess, "slots": placed_slots})
        day_idx += 1
    return assignments

# -------------------------
# Seating allocation per session
# -------------------------
def make_grid(rows, cols):
    return [["" for _ in range(cols)] for __ in range(rows)]

def allocate_seating_for_session(placed_slots, rooms_df, invigators):
    parent_groups = {}
    for slot in placed_slots:
        parent = slot["slot_key"]
        raw = slot.get("slot_raw", parent)
        base = raw.split("_Y")[0] if "_Y" in raw else (raw.split("_")[0] if "_" in raw else raw)
        if slot.get("merged_flag", False):
            items = [{"label_prefix": parent, "remaining": int(slot["students"])}]
        else:
            items = []
            for rr in sorted(slot["courses"], key=lambda x: x["DIVISION"]):
                div = rr["DIVISION"]
                items.append({"label_prefix": f"{parent}_{div}", "remaining": int(rr["NO_STUDENTS"])})
        parent_groups[parent] = {"base": base, "items": items}

    active_items = []
    for parent, v in parent_groups.items():
        for it in v["items"]:
            active_items.append({"parent": parent, "base": v["base"], "label_prefix": it["label_prefix"], "remaining": it["remaining"]})

    rooms_sorted = rooms_df.sort_values(by="Seating Capacity", ascending=False)
    largest_room_cap = rooms_sorted["Seating Capacity"].max() if not rooms_sorted.empty else 0

    rooms = []
    inv_pool = invigators.copy() if invigators else []
    for _, r in rooms_sorted.iterrows():
        cap = int(r["Seating Capacity"])
        rows = 6
        cols = max(1, math.ceil(cap / rows))
        usable = cap // 2

        invs_needed = 2 if cap == largest_room_cap else 1
        alloc_invs = []
        if inv_pool and invs_needed > 0:
            take = min(invs_needed, len(inv_pool))
            for _ in range(take):
                alloc_invs.append(inv_pool.pop(0))
        rooms.append({
            "name": str(r["Room"]),
            "capacity": cap,
            "rows": rows,
            "cols": cols,
            "usable": usable,
            "grid": make_grid(rows, cols),
            "invigilators": alloc_invs
        })

    placed_counters = defaultdict(int)
    for room in rooms:
        if not any(it["remaining"] > 0 for it in active_items):
            break
        rows = room["rows"]
        cols = room["cols"]
        usable = room["usable"]
        placed_in_room = 0
        col_idx = 0
        last_slot_base = None
        while placed_in_room < usable and any(it["remaining"] > 0 for it in active_items) and col_idx < cols:
            chosen_idx = None
            for idx, it in enumerate(active_items):
                if it["remaining"] <= 0:
                    continue
                if last_slot_base is None or it["base"] != last_slot_base:
                    chosen_idx = idx
                    break
            if chosen_idx is None:
                for idx, it in enumerate(active_items):
                    if it["remaining"] > 0:
                        chosen_idx = idx
                        break
                if chosen_idx is None:
                    break
            chosen = active_items[chosen_idx]
            for r_in in range(rows):
                if placed_in_room >= usable:
                    break
                if chosen["remaining"] <= 0:
                    room["grid"][r_in][col_idx] = ""
                    continue
                placed_counters[chosen["label_prefix"]] += 1
                labnum = placed_counters[chosen["label_prefix"]]
                room["grid"][r_in][col_idx] = f"{chosen['label_prefix']}-{labnum}"
                chosen["remaining"] -= 1
                placed_in_room += 1
            last_slot_base = chosen["base"]
            col_idx += 1

    # Fallback fill for leftover
    for room in rooms:
        rows = room["rows"]
        cols = room["cols"]
        usable = room["usable"]
        already = sum(1 for r in range(rows) for c in range(cols) if room["grid"][r][c])
        if already >= usable:
            continue
        for c in range(cols):
            for r_in in range(rows):
                if already >= usable:
                    break
                if room["grid"][r_in][c]:
                    continue
                for it in active_items:
                    if it["remaining"] > 0:
                        placed_counters[it["label_prefix"]] += 1
                        labnum = placed_counters[it["label_prefix"]]
                        room["grid"][r_in][c] = f"{it['label_prefix']}-{labnum}"
                        it["remaining"] -= 1
                        already += 1
                        break

    # ------------------------------------------------------------
    # REDISTRIBUTE INVIGILATORS — DESCENDING CAPACITY LOGIC
    # ------------------------------------------------------------
    rooms_with_students = []
    rooms_empty = []

    for room in rooms:
        occupied = any(room["grid"][r][c] for r in range(room["rows"]) for c in range(room["cols"]))
        if occupied:
            rooms_with_students.append(room)
        else:
            rooms_empty.append(room)

    free_invigs = []
    for room in rooms_empty:
        free_invigs.extend(room["invigilators"])
        room["invigilators"] = []

    free_invigs.extend(inv_pool)

    for room in rooms_with_students:
        if len(room["invigilators"]) > 2:
            extra = room["invigilators"][2:]
            free_invigs.extend(extra)
            room["invigilators"] = room["invigilators"][:2]

    for room in rooms_with_students:
        if len(room["invigilators"]) == 0 and free_invigs:
            room["invigilators"].append(free_invigs.pop(0))

    rooms_sorted_cap_desc = sorted(rooms_with_students, key=lambda x: x["capacity"], reverse=True)

    for room in rooms_sorted_cap_desc:
        if not free_invigs:
            break
        if len(room["invigilators"]) == 1:
            room["invigilators"].append(free_invigs.pop(0))

    for room in rooms_sorted_cap_desc:
        if not free_invigs:
            break
        if len(room["invigilators"]) < 2:
            room["invigilators"].append(free_invigs.pop(0))

    return rooms

# -------------------------
# Write seating Excel (per-day) - will be called for each half saving into given out_dir
# -------------------------
def write_seating_excel(day_idx, rooms_alloc_fn, rooms_alloc_an, day_slots_fn, day_slots_an, df_courses, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    outpath = out_dir / f"Day_{day_idx}.xlsx"
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for sess, rooms_alloc in [("FN", rooms_alloc_fn), ("AN", rooms_alloc_an)]:
        ws = wb.create_sheet(sess)
        row_cursor = 1
        for room in rooms_alloc:
            ws.cell(row=row_cursor, column=1, value=f"Room: {room['name']}")
            inv_display = ", ".join([inv_display_from_key(k) for k in room.get('invigilators', [])])
            ws.cell(row=row_cursor, column=3, value=f"Invigilators: {inv_display}")
            row_cursor += 1
            for r in range(room["rows"]):
                for c in range(room["cols"]):
                    val = room["grid"][r][c] or ""
                    cell = ws.cell(row=row_cursor + r, column=c + 1, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
            row_cursor += room["rows"] + 2
        max_cols = max((room["cols"] for room in rooms_alloc), default=1)
        for c in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

    # -------------------------
    # REFERENCE sheet
    # -------------------------
    ws_ref = wb.create_sheet("REFERENCE")
    headers = ["ELECTIVE OR NOT", "FULLSEM OR HALFSEM", "SLOT NAME", "COURSE CODE", "COURSE TITLE", "DIVISION", "SESSION"]
    ws_ref.append(headers)

    session_map = {}
    for s in day_slots_fn:
        session_map[s["slot_key"]] = "FN"
    for s in day_slots_an:
        session_map[s["slot_key"]] = "AN"

    day_slot_keys = set(session_map.keys())
    if day_slot_keys:
        df_day = df_courses[df_courses["SLOT"].isin(day_slot_keys)].copy()
    else:
        df_day = df_courses.iloc[0:0].copy()

    color_palette = [
        "FFCCCC", "CCFFCC", "CCCCFF", "FFF2CC", "FFD9E6", "E6FFCC", "CCE5FF", "E6CCFF",
        "FFE5CC", "CCFFF2", "F0E68C", "E0FFFF"
    ]
    divisions_list = sorted(df_day["DIVISION"].unique()) if not df_day.empty else []
    div_color_map = {div: color_palette[i % len(color_palette)] for i, div in enumerate(divisions_list)}

    for _, r in df_day.iterrows():
        session = session_map.get(r["SLOT"], "")
        row_vals = [
            r.get("ELECTIVE", ""),
            r.get("FULLSEM_TYPE", ""),
            r.get("SLOT_RAW", ""),
            r.get("COURSE_CODE", ""),
            r.get("COURSE_TITLE", ""),
            r.get("DIVISION", ""),
            session
        ]
        ws_ref.append(row_vals)
        row_idx = ws_ref.max_row
        fill_color = div_color_map.get(r.get("DIVISION", ""), color_palette[0])
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            ws_ref.cell(row=row_idx, column=col_idx).fill = fill

    for col in ws_ref.columns:
        first = col[0]
        try:
            ws_ref.column_dimensions[first.column_letter].width = 25
        except Exception:
            pass
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)

    wb.save(outpath)
    print(f"Wrote seating for Day {day_idx}: {outpath}")

# -------------------------
# Timetable builder (per-half)
# -------------------------
def build_timetable_from_assignments(df_courses, assignments, outpath):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for sess in ("FN", "AN"):
        ws = wb.create_sheet(sess)
        ws.append(["Day", "Slots"])
        for alloc in assignments:
            if alloc["session"] != sess:
                continue
            names = [s["slot_key"] for s in alloc["slots"]]
            ws.append([alloc["day"], ", ".join(names)])
    color_palette = [
        "FFCCCC", "CCFFCC", "CCCCFF", "FFF2CC", "FFD9E6", "E6FFCC", "CCE5FF", "E6CCFF",
        "FFE5CC", "CCFFF2", "F0E68C", "E0FFFF"
    ]
    for year in sorted(df_courses["YEAR"].unique()):
        ws_ref = wb.create_sheet(f"Reference_{year}")
        headers = ["YEAR", "DIVISION", "ELECTIVE", "FULLSEM/HALFSEM", "SLOT NAME", "COURSE CODE", "COURSE TITLE"]
        ws_ref.append(headers)
        ref_df = df_courses[df_courses["YEAR"] == year][
            ["YEAR", "DIVISION", "ELECTIVE", "FULLSEM_TYPE", "SLOT", "COURSE_CODE", "COURSE_TITLE"]
        ].drop_duplicates()
        divisions_list = sorted(ref_df["DIVISION"].unique())
        div_color_map = {div: color_palette[i % len(color_palette)] for i, div in enumerate(divisions_list)}
        for _, r in ref_df.iterrows():
            row_vals = [
                r["YEAR"], r["DIVISION"], r["ELECTIVE"],
                r["FULLSEM_TYPE"], r["SLOT"], r["COURSE_CODE"], r["COURSE_TITLE"]
            ]
            ws_ref.append(row_vals)
            row_idx = ws_ref.max_row
            fill_color = div_color_map.get(r["DIVISION"], color_palette[0])
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws_ref.cell(row=row_idx, column=col_idx).fill = fill
    for sh in wb.sheetnames:
        ws = wb[sh]
        for col in ws.columns:
            first = col[0]
            try:
                ws.column_dimensions[first.column_letter].width = 25
            except Exception:
                pass
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
    Path(outpath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(outpath)
    print(f"Wrote timetable: {outpath}")

# -------------------------
# Build invigilator schedules workbook (per-half)
# -------------------------
def write_invigilator_schedules(invigilator_df, invig_assignments, outpath):
    out_dir = Path(outpath).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet1: copy of provided list
    ws_list = wb.create_sheet("Invigilators_List")
    for c_idx, col in enumerate(invigilator_df.columns.tolist(), start=1):
        ws_list.cell(row=1, column=c_idx, value=col)
    for r_idx, row in invigilator_df.reset_index(drop=True).iterrows():
        for c_idx, col in enumerate(invigilator_df.columns.tolist(), start=1):
            val = row[col]
            try:
                if c_idx == 1:
                    val_int = int(float(val))
                    ws_list.cell(row=r_idx+2, column=c_idx, value=val_int)
                else:
                    ws_list.cell(row=r_idx+2, column=c_idx, value=val)
            except Exception:
                ws_list.cell(row=r_idx+2, column=c_idx, value=val)

    # Per-invigilator sheets
    for _, row in invigilator_df.iterrows():
        num = str(row.iloc[0]).strip()
        name = str(row.iloc[1]).strip() if len(row) > 1 else ""
        sheet_name = f"INVIGILATOR_{num}".upper()[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append(["Day", "Session", "Room"])
        key = inv_key(num, name)
        duties = invig_assignments.get(key, [])
        duties_sorted = sorted(duties, key=lambda x: (x["day"], 0 if x["session"] == "FN" else 1))
        for d in duties_sorted:
            ws.append([d["day"], d["session"], d["room"]])
        for col in ws.columns:
            first = col[0]
            try:
                ws.column_dimensions[first.column_letter].width = 18
            except Exception:
                pass
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)

    wb.save(outpath)
    print(f"Wrote invigilator schedules: {outpath}")

# -------------------------
# Run full generation for a half (keeps algorithm exactly as original)
# -------------------------
def run_half(half_name, courses_df_half, rooms_df, inv_copy_df):
    """
    half_name: "FIRSTHALF" or "SECONDHALF"
    courses_df_half: DataFrame of courses for that half only
    rooms_df: rooms DataFrame
    inv_copy_df: DataFrame (first two columns) from invigilators input
    """
    root_out = Path("EXAM_OUTPUT") / half_name
    seating_out_dir = root_out / "seating_arrangements"
    Path(seating_out_dir).mkdir(parents=True, exist_ok=True)

    # Build invigilator list (preserve order, unique)
    invig_list = []
    invig_seen = set()
    for _, r in inv_copy_df.iterrows():
        num = str(r.iloc[0]).strip()
        name = str(r.iloc[1]).strip() if len(r) > 1 else ""
        if not num:
            continue
        key = inv_key(num, name)
        if key in invig_seen:
            continue
        invig_seen.add(key)
        invig_list.append(key)

    # 1) Slot allocation -> assignments for this half
    assignments = allocate_slots_by_seating_capacity(courses_df_half.copy(), rooms_df)

    # 2) Build timetable file for this half
    timetable_path = root_out / f"{half_name.lower()}_timetable.xlsx"
    build_timetable_from_assignments(courses_df_half, assignments, str(timetable_path))

    # 3) Prepare invigilator assignment mapping for this half
    invig_assignments = defaultdict(list)

    # Calculate total days for this half only (some assign entries may have empty slots)
    total_days = max([alloc["day"] for alloc in assignments]) if assignments else 0

    for day_idx in range(1, total_days + 1):
        # Random split invigilators each day (same logic as original: random shuffle, half/half, extra to AN)
        invig_random = invig_list.copy()
        random.shuffle(invig_random)
        half_invig = len(invig_random) // 2
        invig_fn = invig_random[:half_invig]
        invig_an = invig_random[half_invig:]

        # Extract slots for this half/day/session explicitly from assignments (this half's assignments)
        day_slots_fn = [s for alloc in assignments if alloc["day"] == day_idx and alloc["session"] == "FN" for s in alloc["slots"]]
        day_slots_an = [s for alloc in assignments if alloc["day"] == day_idx and alloc["session"] == "AN" for s in alloc["slots"]]

        # Allocate seating and invigilators for FN and AN
        rooms_alloc_fn = allocate_seating_for_session(day_slots_fn, rooms_df, invig_fn)
        rooms_alloc_an = allocate_seating_for_session(day_slots_an, rooms_df, invig_an)

        # Record invigilator duties
        for room in rooms_alloc_fn:
            for ik in room.get("invigilators", []):
                invig_assignments[ik].append({"day": day_idx, "session": "FN", "room": room["name"]})
        for room in rooms_alloc_an:
            for ik in room.get("invigilators", []):
                invig_assignments[ik].append({"day": day_idx, "session": "AN", "room": room["name"]})

        # Write per-day seating file into this half's folder
        write_seating_excel(day_idx, rooms_alloc_fn, rooms_alloc_an, day_slots_fn, day_slots_an, courses_df_half, seating_out_dir)

    # After all days, write invigilator schedules into this half folder
    inv_sched_path = root_out / "Invigilator_Schedules.xlsx"
    write_invigilator_schedules(inv_copy_df, invig_assignments, str(inv_sched_path))

    print(f"Completed generation for {half_name}. Outputs in: {root_out}")

# -------------------------
# Main
# -------------------------
def main():
    # Load master courses from hardcoded divisions
    df_courses = load_courses(divisions)
    if df_courses.empty:
        print("No courses found. Exiting.")
        return

    first_half_df, second_half_df = split_half(df_courses)

    # Load rooms and invigilators
    rooms_df = pd.read_excel(rooms_path, engine="openpyxl")
    rooms_df.columns = [str(c).strip() for c in rooms_df.columns]
    if "Room" not in rooms_df.columns or "Seating Capacity" not in rooms_df.columns:
        raise ValueError("Rooms file must contain 'Room' and 'Seating Capacity' columns")

    inv_df = pd.read_excel(invig_path, engine="openpyxl", dtype=str)
    inv_df.columns = [str(c).strip() for c in inv_df.columns]
    if inv_df.shape[1] < 1:
        raise ValueError("Invigilator file must have at least one column (NUMBER). Preferably two: NUMBER and NAME.")
    # Keep first two columns as provided
    inv_copy_df = inv_df.iloc[:, :2].copy()
    # Ensure column names are present (we don't force specific names)
    if inv_copy_df.shape[1] == 1:
        inv_copy_df.columns = [inv_copy_df.columns[0]]
    else:
        inv_copy_df.columns = [inv_copy_df.columns[0], inv_copy_df.columns[1]]

    # Create EXAM_OUTPUT root
    Path("EXAM_OUTPUT").mkdir(exist_ok=True)

    # Run FIRSTHALF
    print("\n=== Generating FIRSTHALF ===")
    run_half("FIRSTHALF", first_half_df, rooms_df, inv_copy_df)

    # Run SECONDHALF
    print("\n=== Generating SECONDHALF ===")
    run_half("SECONDHALF", second_half_df, rooms_df, inv_copy_df)

    print("\nAll done. Check EXAM_OUTPUT/FIRSTHALF and EXAM_OUTPUT/SECONDHALF for results.")

if __name__ == "__main__":
    main()
